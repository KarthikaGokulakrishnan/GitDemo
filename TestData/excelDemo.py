import openpyxl

book = openpyxl.load_workbook("C:\\Users\\karthika\\Documents\\PythonDemo.xlsx")
sheet = book.active
Dict={}
cell = sheet.cell(row=1,column=2)  #To read the Value
print(cell.value)
sheet.cell(row=2,column=2).value="Test@123"    #To write into the file
print(sheet.cell(row=2,column=2).value)
print(sheet.max_row)     # to find the  amximum rows and column
print(sheet.max_column)
print(sheet['A2'].value)
for i in range(1,4):  #to print all the values in the sheet   to get rows
    if sheet.cell(row=i,column=1).value =="TestCase2":
        for j in range(1,sheet.max_column+1):  # to get columns
            # Dict["email"]:"kathika@yopmail.com"
            Dict[sheet.cell(row=1,column=j).value]=sheet.cell(row=i,column=j).value
print(Dict)