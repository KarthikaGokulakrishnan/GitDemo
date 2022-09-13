import openpyxl


class HomePageData:
    test_homepage_data = [{"email":"karthi@yopmail.com","Password":"Test@123","Name":"karthika","Gender":"Female"},{"email":"gokul@yopmail.com","Password":"Test@123","Name":"gokul","Gender":"Male"}]
    @staticmethod
    def getData(Test_Case_Name):
        Dict = {}
        book = openpyxl.load_workbook("C:\\Users\\karthika\\Documents\\PythonDemo.xlsx")
        sheet = book.active
        for i in range(1, 4):  # to print all the values in the sheet   to get rows
            if sheet.cell(row=i, column=1).value == Test_Case_Name:
                for j in range(1, sheet.max_column + 1):  # to get columns
                    # Dict["email"]:"kathika@yopmail.com"
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return[Dict]